import pandas as pd
from CreateTransactionModel import (
    BankDetails,
    DateFormat,
)
import xlrd
import openpyxl as xl
import pathlib
import numpy as np
from tkinter.filedialog import askopenfilename


def processfile(path, bankid):
    extension = pathlib.Path(path).suffix
    if extension == ".xlsx":
        engine = "openpyxl"
    elif extension == ".xls":
        engine = "xlrd"

    bank_dets = BankDetails.query.filter_by(bank_id=bankid).first()
    sht_last_row = {}

    if engine == "openpyxl":
        wb = xl.load_workbook(path, read_only=False)
        for sheet in wb:
            f = 0
            for row in sheet.iter_rows(
                min_row=bank_dets.start_row + 1,
                min_col=bank_dets.val_date_col,
                max_col=bank_dets.val_date_col,
            ):
                for cell in row:
                    if cell.value == None or cell.value == "":
                        if (
                            sheet.cell(cell.row + 1, bank_dets.val_date_col).value
                            == None
                            or sheet.cell(cell.row + 1, bank_dets.val_date_col).value
                            == ""
                        ):
                            sht_last_row[sheet.title] = cell.row - 1
                            f = 1
                            break
                if f == 1:
                    break
        wb.close()
    elif engine == "xlrd":
        wb = xlrd.open_workbook(path)
        for sheet_name in wb.sheet_names():
            sht = wb.sheet_by_name(sheet_name)
            f = 0
            for row in range(bank_dets.start_row, sht.nrows):
                for col in range(bank_dets.val_date_col, bank_dets.val_date_col + 1):
                    if (
                        sht.cell(row, col - 1).value == None
                        or sht.cell(row, col - 1).value == ""
                    ):
                        if (
                            sht.cell(row + 1, col - 1).value == None
                            or sht.cell(row + 1, col - 1).value == ""
                        ):
                            f = 1
                            sht_last_row[sheet_name] = row
                            break
                if f == 1:
                    break

    extract_cols = [
        bank_dets.val_date_col - 1,
        bank_dets.txn_date_col - 1,
        bank_dets.chq_no_col - 1,
        bank_dets.txn_rmrk_col - 1,
        bank_dets.with_amt_col - 1,
        bank_dets.crdt_amt_col - 1,
        bank_dets.bal_col - 1,
    ]
    col_names = [
        "value_date",
        "txn_date",
        "cheque_no",
        "txn_remarks",
        "withdrawal_amt",
        "deposit_amt",
        "balance",
    ]
    final_df = pd.DataFrame(columns=col_names)
    sheets = pd.ExcelFile(path, engine=engine)
    for sheet in sheets.sheet_names:
        no_rows = sht_last_row[sheet] - bank_dets.start_row
        df = pd.DataFrame()
        for col_name, col_pos in zip(col_names, extract_cols):
            temp_df = pd.read_excel(
                path,
                sheet_name=sheet,
                usecols=[col_pos],
                skiprows=bank_dets.start_row - 2,
                names=[col_name],
                nrows=no_rows + 1,
            )
            df[col_name] = temp_df

        final_df = pd.concat([final_df, df])
    py_format = DateFormat.query.filter_by(date_id=bank_dets.date_id).first()

    final_df = final_df.dropna(subset="txn_date")

    if not (np.issubdtype(final_df["value_date"].dtype, np.datetime64)):
        final_df["value_date"] = pd.to_datetime(
            final_df["value_date"], format=py_format.py_date
        )
    if not (np.issubdtype(final_df["txn_date"].dtype, np.datetime64)):
        final_df["txn_date"] = pd.to_datetime(
            final_df["txn_date"], format=py_format.py_date
        )

    final_df["withdrawal_amt"] = pd.to_numeric(
        final_df["withdrawal_amt"].astype(str).str.strip().str.replace(",", "")
    )
    final_df["deposit_amt"] = pd.to_numeric(
        final_df["deposit_amt"].astype(str).str.strip().str.replace(",", "")
    )
    final_df["balance"] = pd.to_numeric(
        final_df["balance"].astype(str).str.strip().str.replace(",", "")
    )
    final_df["txn_remarks"] = final_df["txn_remarks"].str.strip()
    final_df["cheque_no"] = final_df["cheque_no"].str.strip()

    final_df = final_df.replace(np.nan, 0)

    return final_df


if __name__ == "__main__":
    ipfilePath = askopenfilename(title="Select Input File")
    bank_id = input("Enter Bank Id: ")
    try:
        int(bank_id)
    except:
        print("Please enter integers only")
    else:
        print(processfile(path=ipfilePath, bankid=bank_id))
